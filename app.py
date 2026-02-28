"""
Mirraw Invoice Extractor
========================
Extracts invoice data from Aashirwad Garments → Mirraw PDF invoices (inside a ZIP),
splits into Registered / Unregistered tabs, flags wrong-seller invoices,
and exports a formatted Excel report.

Run locally:
    pip install streamlit pdfplumber openpyxl pandas
    streamlit run app.py

Deploy: push to GitHub → share.streamlit.io → select repo → main file: app.py
"""

import io
import re
import zipfile

import openpyxl
import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Mirraw Invoice Extractor",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
VALID_SELLER_GSTIN = "08ARNPK0658G1ZL"
VALID_SELLER_NAME  = "AASHIRWAD GARMENTS"

DISPLAY_COLS = [
    "Invoice No", "Invoice Date", "Purchase Order No",
    "Seller Name", "Seller GSTIN",
    "Buyer Name", "Buyer GSTIN", "Registration Type",
    "Place of Supply", "SKU Name", "SKU Code", "Size", "HSN Code",
    "Qty", "Taxable Amount", "GST Rate (%)", "GST Amount (IGST)", "Total Amount",
]

WRONG_SELLER_COLS = [
    "Invoice No", "Invoice Date", "Seller Name", "Seller GSTIN",
    "Buyer Name", "Registration Type", "SKU Name", "SKU Code",
    "Taxable Amount", "Total Amount", "Wrong Seller Reason",
]

NUMERIC_FIELDS = {"Qty", "Taxable Amount", "GST Rate (%)", "GST Amount (IGST)", "Total Amount"}

# ─────────────────────────────────────────────────────────────────────────────
#  CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
.stApp { background-color: #F0F4FF; }
.main-title {
    font-size: 2.2rem; font-weight: 800;
    color: #1A237E; margin-bottom: 4px;
}
.sub-title { font-size: 1rem; color: #5C6BC0; margin-bottom: 20px; }
.metric-card {
    background: white; border-radius: 12px;
    padding: 14px 18px; box-shadow: 0 2px 8px rgba(0,0,0,0.07);
    margin-bottom: 10px;
}
.metric-label {
    font-size: 0.72rem; color: #888;
    font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em;
}
.metric-value { font-size: 1.5rem; font-weight: 800; margin-top: 2px; }
.section-header {
    font-size: 1.05rem; font-weight: 700; color: #1A237E;
    margin: 18px 0 10px; padding-bottom: 6px;
    border-bottom: 2px solid #C5CAE9;
}
.warn-box {
    background: #FFF3E0; border-left: 5px solid #E65100;
    border-radius: 8px; padding: 14px 18px; margin-bottom: 14px;
}
.warn-box b { color: #BF360C; }
.footer {
    text-align: center; color: #9E9E9E; font-size: 0.8rem;
    margin-top: 40px; padding: 16px;
    border-top: 1px solid #E0E0E0;
}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _to_float(val) -> float:
    try:
        return float(val) if val else 0.0
    except (ValueError, TypeError):
        return 0.0


def _parse_nums(text: str) -> list:
    return re.findall(r"\d+(?:\.\d+)?", text)


def _fmt_inr(val) -> str:
    """Format a value as ₹ string, or return as-is if not numeric."""
    try:
        if val == "" or val is None:
            return ""
        return f"₹{float(val):,.2f}"
    except (ValueError, TypeError):
        return str(val)


def _highlight_reg(val):
    if val == "Registered":
        return "background-color:#E8F5E9; color:#1B5E20; font-weight:bold"
    if val == "Unregistered":
        return "background-color:#FFF3E0; color:#BF360C; font-weight:bold"
    return ""


def _highlight_wrong(val):
    return "background-color:#FFEBEE; color:#B71C1C; font-weight:bold"


def render_metric(col, label, value, color):
    with col:
        st.markdown(f"""
        <div class="metric-card" style="border-left:5px solid {color}">
            <div class="metric-label">{label}</div>
            <div class="metric-value" style="color:{color}">{value}</div>
        </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
#  CORE EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────
def extract_from_pdf(pdf_bytes: bytes) -> list:
    """
    Extract invoice line items from a single PDF.
    Adds a 'Wrong Seller' flag if seller GSTIN ≠ VALID_SELLER_GSTIN.
    """
    results = []
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            all_lines = []
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text() or ""
                full_text += text + "\n"
                all_lines.extend(text.split("\n"))

        # Must contain HSN table
        if "HSN_CODE" not in full_text and "HSN CODE" not in full_text:
            return results

        # ── Seller ────────────────────────────────────────────────────────────
        # Seller name: first non-empty line (or after "Name:")
        lines_clean = [l.strip() for l in full_text.split("\n") if l.strip()]
        first_line = lines_clean[0] if lines_clean else ""
        if first_line.startswith("Name:"):
            seller_name = first_line.replace("Name:", "").strip()
        elif first_line in ("AASHIRWAD GARMENTS", "YASH GALLERY"):
            seller_name = first_line
        else:
            # Try to find it via "Sold By" or "Name:" pattern
            m = re.search(r"(?:Sold By|Name)\s*[:\n]\s*([A-Z][^\n,]{3,40})", full_text)
            seller_name = m.group(1).strip() if m else first_line

        m = re.search(r"GST\s*no\s*[:\s]+([A-Z0-9]{15})", full_text)
        if not m:
            m = re.search(r"GSTIN\s*[:\s]+([A-Z0-9]{15})", full_text)
        seller_gstin = m.group(1).strip() if m else ""

        # ── Wrong seller check ─────────────────────────────────────────────────
        wrong_seller = False
        wrong_reason = ""
        if seller_gstin and seller_gstin != VALID_SELLER_GSTIN:
            wrong_seller = True
            wrong_reason = f"Seller GSTIN is {seller_gstin} (expected {VALID_SELLER_GSTIN})"
        elif seller_gstin == "" and "AASHIRWAD" not in full_text.upper():
            wrong_seller = True
            wrong_reason = "Seller GSTIN missing and seller name not Aashirwad Garments"

        # ── Buyer & Registration Type ──────────────────────────────────────────
        m = re.search(r"GST\s*NUMBER\s*:\s*([A-Z0-9]{15})", full_text)
        if m:
            buyer_gstin = m.group(1)
            buyer_name = "Mirraw Online Services Pvt Ltd."
            registration_type = "Registered"
        else:
            buyer_gstin = "N/A"
            m2 = re.search(r"Name:\s*(.+?)(?:\s{2,}|Address:|Invoice)", full_text)
            buyer_name = m2.group(1).strip() if m2 else "Individual / Unregistered"
            registration_type = "Unregistered"

        # ── Invoice header ─────────────────────────────────────────────────────
        m = re.search(r"Invoice No[:\s]+([\w/]+)", full_text)
        invoice_no = m.group(1).strip() if m else ""

        m = re.search(r"Invoice Date\s*[:\s]+([\d/\s]+)", full_text)
        invoice_date = m.group(1).strip().replace(" ", "") if m else ""

        m = re.search(r"Purchase Order No[:\s]+([\w]+)", full_text)
        purchase_order_no = m.group(1).strip() if m else ""

        m = re.search(r"Place of Supply\s*:\s*([^\n]+)", full_text)
        place_of_supply = m.group(1).strip() if m else ""

        m = re.search(r"State Code\s*:\s*(\d+)", full_text)
        state_code = m.group(1).strip() if m else ""

        # ── Line items ─────────────────────────────────────────────────────────
        i = 0
        while i < len(all_lines):
            line = all_lines[i]

            if "[sku:" not in line.lower():
                i += 1
                continue

            # Pattern A: SKU closed inline before HSN  "Tops [sku: ABC123] 61149090 1 …"
            m_inline = re.search(r"\[sku:\s*([A-Za-z0-9_\-]+)\]\s*(\d{8})", line)

            if m_inline:
                sku_code = m_inline.group(1).strip()
                hsn      = m_inline.group(2)
                m2       = re.match(r"^(.+?)\s*\[sku:", line)
                sku_name = m2.group(1).strip() if m2 else ""
                size     = ""
                if i + 1 < len(all_lines):
                    m3 = re.search(r"\[size[:\s]+([^\]]+)\]", all_lines[i + 1], re.IGNORECASE)
                    if m3:
                        size = m3.group(1).strip()
                nums = _parse_nums(line[line.find(hsn) + len(hsn):])

            else:
                # Pattern B: SKU code on next line  "Plus size kurtis [sku: 61149090 1 …"
                #                                   "SKUCODE] [size: 3XL]"
                m_hsn = re.search(r"\b(\d{8})\b", line)
                hsn   = m_hsn.group(1) if m_hsn else ""

                next1 = all_lines[i + 1] if i + 1 < len(all_lines) else ""
                next2 = all_lines[i + 2] if i + 2 < len(all_lines) else ""

                m2       = re.match(r"^([A-Za-z0-9_\-]+)\]", next1.strip())
                sku_code = m2.group(1).strip() if m2 else ""
                m3       = re.match(r"^(.+?)\s*\[sku:", line)
                sku_name = m3.group(1).strip() if m3 else ""
                m4       = re.search(r"\[size[:\s]+([^\]]+)\]", next1 + " " + next2, re.IGNORECASE)
                size     = m4.group(1).strip() if m4 else ""
                nums     = _parse_nums(line[line.find(hsn) + len(hsn):] if hsn else line)

            if not sku_code and not hsn:
                i += 1
                continue

            # Nums: QTY VALUE [DISC] NET TAXABLE IGST_RATE IGST_AMT TOTAL
            qty = taxable = igst_rate = igst_amt = total = ""
            if len(nums) >= 8:
                qty, taxable, igst_rate, igst_amt, total = nums[0], nums[4], nums[5], nums[6], nums[7]
            elif len(nums) >= 7:
                qty, taxable, igst_rate, igst_amt, total = nums[0], nums[3], nums[4], nums[5], nums[6]
            elif len(nums) >= 6:
                qty, taxable, igst_rate, igst_amt = nums[0], nums[3], nums[4], nums[5]
            elif len(nums) >= 1:
                qty = nums[0]

            results.append({
                "Invoice No":         invoice_no,
                "Invoice Date":       invoice_date,
                "Purchase Order No":  purchase_order_no,
                "Seller Name":        seller_name,
                "Seller GSTIN":       seller_gstin,
                "Buyer Name":         buyer_name,
                "Buyer GSTIN":        buyer_gstin,
                "Registration Type":  registration_type,
                "Place of Supply":    place_of_supply,
                "State Code":         state_code,
                "SKU Name":           sku_name,
                "SKU Code":           sku_code,
                "Size":               size,
                "HSN Code":           hsn,
                "Qty":                qty,
                "Taxable Amount":     taxable,
                "GST Rate (%)":       igst_rate,
                "GST Amount (IGST)":  igst_amt,
                "Total Amount":       total,
                "Wrong Seller":       wrong_seller,
                "Wrong Seller Reason": wrong_reason,
            })
            i += 1

    except Exception:
        pass

    return results


def process_zip(zip_bytes: bytes) -> list:
    """Extract all PDFs from a ZIP and return deduplicated invoice rows."""
    all_data = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
        pdf_names = sorted([n for n in z.namelist() if n.lower().endswith(".pdf")])
        total     = len(pdf_names)
        bar       = st.progress(0, text="Reading invoices…")
        for idx, name in enumerate(pdf_names):
            with z.open(name) as f:
                all_data.extend(extract_from_pdf(f.read()))
            bar.progress((idx + 1) / total, text=f"Processing {idx+1}/{total} PDFs…")
        bar.empty()

    # Deduplicate on (Invoice No, SKU Code)
    seen   = set()
    unique = []
    for row in all_data:
        if not row["SKU Code"] and not row["HSN Code"]:
            continue
        key = (row["Invoice No"], row["SKU Code"])
        if key not in seen:
            seen.add(key)
            unique.append(row)

    unique.sort(key=lambda r: r["Invoice No"])
    return unique


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL BUILDER  (3 sheets: All | Registered | Unregistered)
# ─────────────────────────────────────────────────────────────────────────────
HEADERS = [
    "Sr No", "Invoice No", "Invoice Date", "Purchase Order No",
    "Seller Name", "Seller GSTIN", "Buyer Name", "Buyer GSTIN",
    "Registration Type", "Place of Supply", "State Code",
    "SKU Name", "SKU Code", "Size", "HSN Code",
    "Qty", "Taxable Amount (₹)", "GST Rate (%)", "GST Amount/IGST (₹)", "Total Amount (₹)",
]
COL_WIDTHS = [6, 22, 13, 22, 24, 18, 28, 18, 16, 18, 11, 20, 22, 12, 12, 6, 18, 13, 20, 16]
DATA_FIELDS = [
    "Invoice No", "Invoice Date", "Purchase Order No",
    "Seller Name", "Seller GSTIN", "Buyer Name", "Buyer GSTIN",
    "Registration Type", "Place of Supply", "State Code",
    "SKU Name", "SKU Code", "Size", "HSN Code",
    "Qty", "Taxable Amount", "GST Rate (%)", "GST Amount (IGST)", "Total Amount",
]

_THIN        = Side(style="thin", color="D0D0D0")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_NAVY        = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
_GREEN_HDR   = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
_ORANGE_HDR  = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
_EVEN        = PatternFill(start_color="EEF0FB", end_color="EEF0FB", fill_type="solid")
_ODD         = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_REG_FILL    = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
_UNREG_FILL  = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
# SUM columns: (1-based col index, column letter)
_SUM_COLS = [(16, "P"), (17, "Q"), (19, "S"), (20, "T")]


def _write_sheet(ws, data: list, header_fill):
    """Write headers + data rows + totals row onto a worksheet."""
    ws.row_dimensions[1].height = 40
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    for row_idx, row in enumerate(data, 2):
        ws.row_dimensions[row_idx].height = 18
        row_fill = _EVEN if row_idx % 2 == 0 else _ODD

        # Sr No
        c = ws.cell(row=row_idx, column=1, value=row_idx - 1)
        c.font = Font(name="Arial", size=9); c.fill = row_fill
        c.border = _BORDER; c.alignment = Alignment(horizontal="center", vertical="center")

        for col, field in enumerate(DATA_FIELDS, 2):
            raw = row.get(field, "")
            val = _to_float(raw) if field in NUMERIC_FIELDS and raw else raw

            cell        = ws.cell(row=row_idx, column=col, value=val)
            cell.border = _BORDER

            if field == "Registration Type":
                if val == "Registered":
                    cell.font = Font(name="Arial", size=9, color="1B5E20", bold=True)
                    cell.fill = _REG_FILL
                else:
                    cell.font = Font(name="Arial", size=9, color="BF360C", bold=True)
                    cell.fill = _UNREG_FILL
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif field in {"Taxable Amount", "GST Amount (IGST)", "Total Amount"}:
                cell.font          = Font(name="Arial", size=9)
                cell.fill          = row_fill
                cell.number_format = "#,##0.00"
                cell.alignment     = Alignment(horizontal="right", vertical="center")
            elif field in {"Qty", "GST Rate (%)"}:
                cell.font          = Font(name="Arial", size=9)
                cell.fill          = row_fill
                cell.number_format = "0.00"
                cell.alignment     = Alignment(horizontal="center", vertical="center")
            else:
                cell.font      = Font(name="Arial", size=9)
                cell.fill      = row_fill
                cell.alignment = Alignment(
                    horizontal="center" if field == "Invoice Date" else "left",
                    vertical="center",
                )

    # Totals row
    tr = len(data) + 2
    for c in range(1, len(HEADERS) + 1):
        cell        = ws.cell(row=tr, column=c)
        cell.font   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        cell.fill   = header_fill
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.cell(row=tr, column=1, value="TOTAL")
    for col_idx, col_letter in _SUM_COLS:
        cell               = ws.cell(row=tr, column=col_idx,
                                     value=f"=SUM({col_letter}2:{col_letter}{tr-1})")
        cell.number_format = "#,##0.00"
        cell.alignment     = Alignment(horizontal="right", vertical="center")

    ws.freeze_panes = "A2"


def build_excel(data: list) -> bytes:
    """
    Build a 3-sheet Excel workbook:
      Sheet 1 – All Invoices
      Sheet 2 – Registered only
      Sheet 3 – Unregistered only
    """
    wb = openpyxl.Workbook()

    registered   = [r for r in data if r["Registration Type"] == "Registered"]
    unregistered = [r for r in data if r["Registration Type"] == "Unregistered"]

    ws_all  = wb.active
    ws_all.title = "All Invoices"
    _write_sheet(ws_all, data, _NAVY)

    ws_reg  = wb.create_sheet("Registered")
    _write_sheet(ws_reg, registered, _GREEN_HDR)

    ws_unreg = wb.create_sheet("Unregistered")
    _write_sheet(ws_unreg, unregistered, _ORANGE_HDR)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
#  REUSABLE: render a data tab (filters + table + charts + downloads)
# ─────────────────────────────────────────────────────────────────────────────
def render_data_tab(df_tab: pd.DataFrame, tab_key: str, excel_data: list):
    """Render filter controls, table, charts and download buttons for a subset."""

    if df_tab.empty:
        st.info("No records in this category.")
        return

    # Metrics
    c1, c2, c3, c4 = st.columns(4)
    render_metric(c1, "📄 Invoices",      str(int(df_tab["Invoice No"].nunique())),          "#3F51B5")
    render_metric(c2, "📦 Total Qty",     str(int(df_tab["Qty"].apply(_to_float).sum())),    "#00897B")
    render_metric(c3, "💰 Taxable",       f"₹{df_tab['Taxable Amount'].apply(_to_float).sum():,.2f}", "#F57C00")
    render_metric(c4, "🧾 Total Amount",  f"₹{df_tab['Total Amount'].apply(_to_float).sum():,.2f}",   "#C62828")

    st.markdown("")

    # Filters
    fc1, fc2, fc3 = st.columns(3)
    with fc1:
        inv_f = st.multiselect(
            "Invoice No", options=sorted(df_tab["Invoice No"].unique()),
            default=[], key=f"inv_{tab_key}"
        )
    with fc2:
        sku_f = st.multiselect(
            "SKU Code", options=sorted(df_tab["SKU Code"].unique()),
            default=[], key=f"sku_{tab_key}"
        )
    with fc3:
        date_f = st.multiselect(
            "Invoice Date", options=sorted(df_tab["Invoice Date"].unique()),
            default=[], key=f"date_{tab_key}"
        )

    filtered = df_tab.copy()
    if inv_f:  filtered = filtered[filtered["Invoice No"].isin(inv_f)]
    if sku_f:  filtered = filtered[filtered["SKU Code"].isin(sku_f)]
    if date_f: filtered = filtered[filtered["Invoice Date"].isin(date_f)]

    st.caption(f"Showing **{len(filtered)}** of {len(df_tab)} records")

    # Table
    display_df = filtered[DISPLAY_COLS].reset_index(drop=True)

    def _safe_fmt(v):
        try:
            if v == "" or v is None:
                return ""
            return f"₹{float(v):,.2f}"
        except Exception:
            return str(v)

    styled = (
        display_df.style
        .applymap(_highlight_reg, subset=["Registration Type"])
        .format({
            "Taxable Amount":    _safe_fmt,
            "GST Amount (IGST)": _safe_fmt,
            "Total Amount":      _safe_fmt,
        })
    )
    st.dataframe(styled, use_container_width=True, height=400)

    # Charts
    st.markdown('<div class="section-header">📊 Charts</div>', unsafe_allow_html=True)
    ch1, ch2 = st.columns(2)

    with ch1:
        st.markdown("**Sales by Date (₹)**")
        date_chart = (
            filtered.assign(Amt=filtered["Total Amount"].apply(_to_float))
            .groupby("Invoice Date")["Amt"].sum()
        )
        if not date_chart.empty:
            st.bar_chart(date_chart)

    with ch2:
        st.markdown("**Top SKUs by Amount (₹)**")
        sku_chart = (
            filtered.assign(Amt=filtered["Total Amount"].apply(_to_float))
            .groupby("SKU Code")["Amt"].sum()
            .sort_values(ascending=False)
            .head(10)
        )
        if not sku_chart.empty:
            st.bar_chart(sku_chart)

    # Downloads
    st.markdown('<div class="section-header">⬇️ Download</div>', unsafe_allow_html=True)
    dl1, dl2 = st.columns(2)

    with dl1:
        excel_bytes = build_excel(excel_data)
        st.download_button(
            label="📥 Download Full Excel (3 sheets)",
            data=excel_bytes,
            file_name="Mirraw_Invoices_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
            key=f"xl_{tab_key}",
        )

    with dl2:
        csv_bytes = filtered[DISPLAY_COLS].to_csv(index=False).encode("utf-8")
        st.download_button(
            label="📄 Download Filtered CSV",
            data=csv_bytes,
            file_name=f"Mirraw_{tab_key}_Filtered.csv",
            mime="text/csv",
            use_container_width=True,
            key=f"csv_{tab_key}",
        )


# ─────────────────────────────────────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 🧾 Mirraw Invoice Extractor")
    st.markdown("---")
    st.markdown("""
**How to use:**
1. Upload your **ZIP** file of invoice PDFs
2. Click **Extract Invoices**
3. View data in tabs:
   - 🟢 Registered
   - 🟠 Unregistered
   - ⚠️ Wrong Seller
4. Download Excel / CSV

**Tabs explained:**
- 🟢 **Registered** — Buyer has valid GSTIN
- 🟠 **Unregistered** — Individual buyer, no GSTIN
- ⚠️ **Wrong Seller** — Seller GSTIN ≠ `08ARNPK0658G1ZL`
    """)
    st.markdown("---")
    st.markdown(f"**Expected Seller GSTIN:**  \n`{VALID_SELLER_GSTIN}`")
    st.markdown(f"**Expected Seller Name:**  \n`{VALID_SELLER_NAME}`")
    st.markdown("---")
    st.caption("Python · pdfplumber · openpyxl · Streamlit")


# ─────────────────────────────────────────────────────────────────────────────
#  MAIN UI
# ─────────────────────────────────────────────────────────────────────────────
st.markdown('<div class="main-title">🧾 Mirraw Invoice Extractor</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-title">Extract & Export · Aashirwad Garments → Mirraw Sales Invoices</div>', unsafe_allow_html=True)

uploaded = st.file_uploader(
    "📁 Upload Invoice ZIP file",
    type=["zip"],
    help="ZIP containing Aashirwad Garments PDF invoices",
)

if uploaded:
    st.success(f"✅ Uploaded: **{uploaded.name}** ({uploaded.size / 1024:.1f} KB)")

    if st.button("🚀 Extract Invoices", use_container_width=True, type="primary"):
        with st.spinner("Scanning PDFs and extracting data…"):
            data = process_zip(uploaded.read())

        if not data:
            st.error("❌ No valid invoices found. Please check the ZIP file.")
        else:
            st.session_state["invoice_data"] = data
            wrong_count = sum(1 for r in data if r.get("Wrong Seller"))
            msg = f"✅ Extracted **{len(data)} invoice line items**"
            if wrong_count:
                msg += f" — ⚠️ **{wrong_count} wrong-seller item(s) detected!**"
            st.success(msg)

# ── Results ──────────────────────────────────────────────────────────────────
if "invoice_data" in st.session_state:
    data = st.session_state["invoice_data"]
    df   = pd.DataFrame(data)

    # ── Top-level metrics ─────────────────────────────────────────────────────
    total_invoices  = int(df["Invoice No"].nunique())
    total_qty       = int(df["Qty"].apply(_to_float).sum())
    total_taxable   = df["Taxable Amount"].apply(_to_float).sum()
    total_gst       = df["GST Amount (IGST)"].apply(_to_float).sum()
    total_amount    = df["Total Amount"].apply(_to_float).sum()
    reg_count       = int((df["Registration Type"] == "Registered").sum())
    unreg_count     = int((df["Registration Type"] == "Unregistered").sum())
    wrong_count     = int(df["Wrong Seller"].sum())

    c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
    render_metric(c1, "📄 Total Invoices",  str(total_invoices),          "#3F51B5")
    render_metric(c2, "📦 Total Qty",       str(total_qty),               "#00897B")
    render_metric(c3, "💰 Taxable",         f"₹{total_taxable:,.2f}",     "#F57C00")
    render_metric(c4, "🏛️ GST",            f"₹{total_gst:,.2f}",         "#7B1FA2")
    render_metric(c5, "🧾 Total Amount",    f"₹{total_amount:,.2f}",      "#C62828")
    render_metric(c6, "🟢 Registered",      str(reg_count),               "#2E7D32")
    render_metric(c7, "🟠 Unregistered",    str(unreg_count),             "#BF360C")

    # ── Wrong seller banner ───────────────────────────────────────────────────
    if wrong_count:
        st.markdown(f"""
        <div class="warn-box">
            ⚠️ <b>Warning:</b> {wrong_count} invoice line item(s) have a <b>wrong or unexpected seller</b>.
            These are shown in the <b>⚠️ Wrong Seller</b> tab below. Please verify with Mirraw.
        </div>""", unsafe_allow_html=True)

    st.markdown("---")

    # ── TABS ──────────────────────────────────────────────────────────────────
    tab_labels = [
        f"🟢 Registered ({reg_count})",
        f"🟠 Unregistered ({unreg_count})",
        f"⚠️ Wrong Seller ({wrong_count})",
        "📊 Overview",
    ]
    tab_reg, tab_unreg, tab_wrong, tab_overview = st.tabs(tab_labels)

    df_reg   = df[df["Registration Type"] == "Registered"].copy()
    df_unreg = df[df["Registration Type"] == "Unregistered"].copy()
    df_wrong = df[df["Wrong Seller"] == True].copy()

    # ── Registered tab ────────────────────────────────────────────────────────
    with tab_reg:
        st.markdown('<div class="section-header">🟢 Registered Buyer Invoices</div>', unsafe_allow_html=True)
        render_data_tab(df_reg, "registered", data)

    # ── Unregistered tab ──────────────────────────────────────────────────────
    with tab_unreg:
        st.markdown('<div class="section-header">🟠 Unregistered Buyer Invoices</div>', unsafe_allow_html=True)
        render_data_tab(df_unreg, "unregistered", data)

    # ── Wrong Seller tab ──────────────────────────────────────────────────────
    with tab_wrong:
        st.markdown('<div class="section-header">⚠️ Wrong / Unexpected Seller Invoices</div>', unsafe_allow_html=True)

        if df_wrong.empty:
            st.success("✅ No wrong-seller invoices found. All invoices are from the expected seller.")
        else:
            st.markdown(f"""
            <div class="warn-box">
                Found <b>{len(df_wrong)}</b> line item(s) where the seller is <b>NOT</b>
                <code>{VALID_SELLER_GSTIN}</code> ({VALID_SELLER_NAME}).
                These invoices were shared by Mirraw but belong to a different seller.
                Please contact Mirraw to correct these.
            </div>""", unsafe_allow_html=True)

            # Show wrong seller table with reason
            wrong_display_cols = [c for c in WRONG_SELLER_COLS if c in df_wrong.columns]
            wrong_df = df_wrong[wrong_display_cols].reset_index(drop=True)

            styled_wrong = wrong_df.style.applymap(
                _highlight_wrong, subset=["Seller GSTIN", "Wrong Seller Reason"]
            )
            st.dataframe(styled_wrong, use_container_width=True, height=350)

            # Download wrong seller invoices as CSV
            csv_wrong = df_wrong[wrong_display_cols].to_csv(index=False).encode("utf-8")
            st.download_button(
                label="📄 Download Wrong Seller List (CSV)",
                data=csv_wrong,
                file_name="Mirraw_Wrong_Seller_Invoices.csv",
                mime="text/csv",
                use_container_width=True,
            )

    # ── Overview tab ──────────────────────────────────────────────────────────
    with tab_overview:
        st.markdown('<div class="section-header">📊 Overall Summary</div>', unsafe_allow_html=True)

        ov1, ov2 = st.columns(2)

        with ov1:
            st.markdown("**Daily Sales (₹) — All Invoices**")
            daily = (
                df.assign(Amt=df["Total Amount"].apply(_to_float))
                .groupby("Invoice Date")["Amt"].sum()
            )
            if not daily.empty:
                st.bar_chart(daily)

        with ov2:
            st.markdown("**Registration Type Distribution**")
            reg_series = df["Registration Type"].value_counts()
            reg_chart  = pd.DataFrame({"Count": reg_series.values}, index=reg_series.index)
            reg_chart.index.name = "Type"
            st.bar_chart(reg_chart)

        ov3, ov4 = st.columns(2)

        with ov3:
            st.markdown("**Top 10 SKUs by Revenue (₹)**")
            top_sku = (
                df.assign(Amt=df["Total Amount"].apply(_to_float))
                .groupby("SKU Code")["Amt"].sum()
                .sort_values(ascending=False)
                .head(10)
            )
            if not top_sku.empty:
                st.bar_chart(top_sku)

        with ov4:
            st.markdown("**Sales by Place of Supply (₹)**")
            by_state = (
                df.assign(Amt=df["Total Amount"].apply(_to_float))
                .groupby("Place of Supply")["Amt"].sum()
                .sort_values(ascending=False)
            )
            if not by_state.empty:
                st.bar_chart(by_state)

        # Full download
        st.markdown("---")
        excel_bytes = build_excel(data)
        st.download_button(
            label="📥 Download Full Excel Report (All 3 Sheets)",
            data=excel_bytes,
            file_name="Mirraw_Invoices_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
            key="xl_overview",
        )

else:
    st.markdown("""
    <div style="text-align:center;padding:60px 0;color:#9E9E9E">
        <div style="font-size:4rem">📁</div>
        <div style="font-size:1.2rem;margin-top:12px">Upload a ZIP file to get started</div>
        <div style="font-size:0.9rem;margin-top:8px">Supports Aashirwad Garments → Mirraw PDF invoices</div>
    </div>
    """, unsafe_allow_html=True)

st.markdown(
    '<div class="footer">Mirraw Invoice Extractor · Aashirwad Garments · Streamlit + Python</div>',
    unsafe_allow_html=True,
)
